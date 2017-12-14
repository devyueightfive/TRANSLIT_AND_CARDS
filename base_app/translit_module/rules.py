def get_rules(file="./rules.xlsx", sheetname='rules', debug="N"):
    """fill dictionary with rules and return it
    \n  file -
    \n  sheetname -
    \n  debug - option to print debug messages to console (in case debug = 'Y')
    """
    from . import helpfull_output
    helpfull_output.deb_print(debug, "Welcome to 'get_rules()' function")
    result = {}
    # open excel-'file'
    # use excel library 'openpyxl'
    import openpyxl
    try:
        helpfull_output.deb_print(debug, 'Opening a file ', file)
        wb = openpyxl.load_workbook(file)
        obj_rules_sheet = wb.get_sheet_by_name(sheetname)
    except Exception as wb_ex:
        helpfull_output.deb_print(debug, wb_ex)
    else:
        helpfull_output.deb_print(debug, "Found '", obj_rules_sheet.title, "' sheet")
        helpfull_output.deb_print(debug, "Adding rules..")
        i = 1
        while obj_rules_sheet.cell(row=i, column=1).value is not None:
            result[obj_rules_sheet.cell(row=i, column=1).value] = \
                obj_rules_sheet.cell(row=i, column=2).value
            i += 1
        helpfull_output.deb_print(debug, "Closing the file...")
        wb.close()
    finally:
        helpfull_output.deb_print(debug, "Return result. \nEnd of function.\n\n")
        return result
